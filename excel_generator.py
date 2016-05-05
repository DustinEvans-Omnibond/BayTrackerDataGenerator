#-------------------------------------------------------------------------------
# Name:        excel_generator.py
# Purpose:
#
# Author:      Dustin Evans
#
# Created:     26/04/2016
# Copyright:   (c) Dustin Evans 2016
# Licence:     <your licence>
#-------------------------------------------------------------------------------

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from datetime import datetime, timedelta
from pytz import timezone


#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
def write_data(data_dict, dest_filename, default_timezone):
    wb = Workbook()

    sheet = wb.active
    sheet.title = 'BayTracker Data'

    # Create meta info headers if available
    current_row = 1
    if 'headers' in data_dict:
        meta_headers = data_dict['headers']
        for header in meta_headers:
            cell_coord = 'A' + str(current_row)
            sheet[cell_coord].font = Font(bold=True, size=18)

            _ = sheet.cell(column=1, row=current_row, value=header)
            current_row += 1


    # Create column headers and get cell widths
    col_headers = ['Bay', 'Start Time', 'Service Time', 'Wait Time', 'Total Customer Time', 'Snapshot']
    col_widths = [40, 40, 40, 40, 40, 55]
    for col in range(1, len(col_headers)+1):
        cell_coord = ("%s" % get_column_letter(col)) + str(current_row)
        sheet[cell_coord].font = Font(bold=True, size=18)
        sheet[cell_coord].alignment = Alignment(horizontal='center', vertical='center')
        sheet[cell_coord].fill = PatternFill(fill_type='solid', start_color='e5e5e5e5', end_color='e5e5e5e5')
        sheet.column_dimensions[get_column_letter(col)].width = col_widths[col-1]

        _ = sheet.cell(column=col, row=current_row, value=col_headers[col-1])


    # Write data to sheet
    vehicles = data_dict['vehicles']
    current_row += 1
    for obj in vehicles:
        sheet.row_dimensions[current_row].height = 200

        bay = obj['bay']
        tz = timezone(default_timezone)
        if 'timezone' in obj:
            tz = timezone(obj['timezone'])
        start_time = datetime.fromtimestamp(obj['t_enter'], tz).strftime('%Y/%m/%d %I:%M %p')
        service_time = abs(obj['t_leave'] - obj['t_enter'])
        wait_time = abs(obj['t_enter'] - obj['t_queue_enter'])
        total_customer_time = service_time + wait_time
        snapshot = obj['snapshot']

        col_entries = [bay, start_time, str(timedelta(seconds=service_time)), str(timedelta(seconds=wait_time)), str(timedelta(seconds=total_customer_time)), snapshot]
        for col in xrange(1, len(col_entries)+1):
            cell_coord = ("%s" % get_column_letter(col)) + str(current_row)
            sheet[cell_coord].font = Font(bold=True,size=14)
            sheet[cell_coord].alignment = Alignment(horizontal='center', vertical='center')

            if col == len(col_entries):
                # Snapshot
                img = Image(col_entries[col-1])
                sheet.add_image(img, cell_coord)
            else:
                # Cell text
                _ = sheet.cell(column=col, row=current_row, value=col_entries[col-1])

        current_row += 1


    # Create footers if exists in data_dict
    if 'footers' in data_dict:
        footers = data_dict['footers']
        for footer in footers:
            cell_coord = 'A' + str(current_row)
            sheet[cell_coord].font = Font(bold=True, size=18)

            _ = sheet.cell(column=1, row=current_row, value=footer)
            current_row += 1

    # Write Excell workbook to destination file
    wb.save(filename=dest_filename)